import streamlit as st
import pandas as pd
from datetime import datetime
import tempfile
import os
import re
import json
from openpyxl import load_workbook

st.set_page_config(page_title="Excel不良报表汇总系统", layout="wide")

# ========== 配置文件 ==========
CONFIG_FILE = "usage_stats.json"
ADMIN_PASSWORD = "admin123"  # 管理员密码（输入此密码进入管理员模式）
USER_PASSWORD = "123456"  # 普通用户密码（输入此密码进入用户模式）


# ========== 访问统计功能 ==========
def get_client_ip():
    """获取客户端标识"""
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        ctx = get_script_run_ctx()
        if ctx and ctx.session_id:
            return ctx.session_id[:8]
    except:
        pass
    return "未知"


def load_stats():
    """加载统计数据"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {"visits": [], "total_usage": 0}


def save_stats(stats):
    """保存统计数据"""
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)


def log_usage(action, file_count=0, details=""):
    """记录使用日志"""
    stats = load_stats()

    visit_record = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "session_id": get_client_ip(),
        "action": action,
        "file_count": file_count,
        "details": details
    }

    stats["visits"].append(visit_record)
    stats["total_usage"] = stats.get("total_usage", 0) + 1

    if len(stats["visits"]) > 500:
        stats["visits"] = stats["visits"][-500:]

    save_stats(stats)


def get_stats_summary():
    """获取统计摘要"""
    stats = load_stats()
    visits = stats.get("visits", [])

    if not visits:
        return "暂无使用记录"

    summary_actions = [v for v in visits if v["action"] == "汇总完成"]
    today = datetime.now().strftime("%Y-%m-%d")
    today_visits = [v for v in visits if v["timestamp"].startswith(today)]

    return {
        "total_visits": len(visits),
        "total_summaries": len(summary_actions),
        "today_visits": len(today_visits),
        "recent_visits": visits[-10:][::-1]
    }


# ========== 管理员面板 ==========
def admin_panel():
    """管理员面板"""
    st.title("🔐 管理员面板")

    stats = load_stats()
    summary = get_stats_summary()

    if isinstance(summary, dict):
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("总访问次数", summary["total_visits"])
        with col2:
            st.metric("汇总操作次数", summary["total_summaries"])
        with col3:
            st.metric("今日访问", summary["today_visits"])

        st.divider()
        st.subheader("📋 最近使用记录")

        if summary["recent_visits"]:
            df_logs = pd.DataFrame(summary["recent_visits"])
            st.dataframe(df_logs, use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("📊 每日使用统计")

        daily_stats = {}
        for visit in stats.get("visits", []):
            date = visit["timestamp"][:10]
            if date not in daily_stats:
                daily_stats[date] = {"visits": 0, "summaries": 0}
            daily_stats[date]["visits"] += 1
            if visit["action"] == "汇总完成":
                daily_stats[date]["summaries"] += 1

        if daily_stats:
            df_daily = pd.DataFrame([
                {"日期": date, "访问次数": data["visits"], "汇总次数": data["summaries"]}
                for date, data in sorted(daily_stats.items(), reverse=True)
            ])
            st.dataframe(df_daily, use_container_width=True, hide_index=True)

        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📥 下载完整日志",
                data=json.dumps(stats, ensure_ascii=False, indent=2),
                file_name=f"usage_logs_{datetime.now().strftime('%Y%m%d')}.json",
                mime="application/json",
                use_container_width=True
            )
        with col2:
            if st.button("🏠 返回主页", use_container_width=True):
                st.session_state.admin_mode = False
                st.rerun()
    else:
        st.info(summary)
        if st.button("返回主页"):
            st.session_state.admin_mode = False
            st.rerun()


# ========== 密码保护 ==========
def check_password():
    """验证密码 - 单密码框，自动识别身份"""
    if "password_correct" not in st.session_state:
        st.session_state.password_correct = False
    if "admin_mode" not in st.session_state:
        st.session_state.admin_mode = False

    if st.session_state.password_correct:
        return True

    # 登录界面
    st.markdown("""
    <style>
    .login-container {
        max-width: 400px;
        margin: 100px auto;
        padding: 40px;
        background: white;
        border-radius: 20px;
        box-shadow: 0 10px 40px rgba(0,0,0,0.1);
        text-align: center;
    }
    .login-title {
        font-size: 28px;
        font-weight: bold;
        color: #333;
        margin-bottom: 30px;
    }
    .login-hint {
        font-size: 14px;
        color: #999;
        margin-top: 20px;
    }
    </style>
    """, unsafe_allow_html=True)

    with st.container():
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown('<div class="login-container">', unsafe_allow_html=True)
            st.markdown('<p class="login-title">📊 机台不良报表汇总系统</p>', unsafe_allow_html=True)

            password = st.text_input("请输入访问密码", type="password", placeholder="请输入密码")

            if st.button("登 录", type="primary", use_container_width=True):
                if password == ADMIN_PASSWORD:
                    st.session_state.password_correct = True
                    st.session_state.admin_mode = True
                    log_usage("管理员登录")
                    st.rerun()
                elif password == USER_PASSWORD:
                    st.session_state.password_correct = True
                    st.session_state.admin_mode = False
                    log_usage("用户登录")
                    st.rerun()
                else:
                    st.error("❌ 密码错误，请重试")

            st.markdown('<p class="login-hint">如需访问请联系管理员</p>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

    return False


# ========== 核心功能函数 ==========
def extract_number(s):
    if pd.isna(s) or s == '':
        return 999
    s = str(s).strip()
    numbers = re.findall(r'\d+', s)
    if numbers:
        return int(numbers[0])
    return 999


def normalize_machine_name(machine_str):
    if pd.isna(machine_str) or machine_str == '':
        return ''
    machine_str = str(machine_str).strip()
    if '#' in machine_str:
        return machine_str
    numbers = re.findall(r'\d+', machine_str)
    if numbers:
        return f"{numbers[0]}#"
    return machine_str


def parse_cell_value(cell):
    if cell is None:
        return 0
    if hasattr(cell, 'value'):
        val = cell.value
    else:
        val = cell
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip()
    if val_str.startswith('='):
        numbers = re.findall(r'\d+\.?\d*', val_str)
        return sum(float(n) for n in numbers) if numbers else 0
    try:
        return float(val_str)
    except:
        return 0


def process_files(uploaded_files):
    daily_summary = {}
    machine_data = {}
    employee_data = {}

    progress_bar = st.progress(0)
    status_text = st.empty()

    for i, uploaded_file in enumerate(uploaded_files):
        status_text.text(f"正在处理: {uploaded_file.name} ({i + 1}/{len(uploaded_files)})")

        try:
            file_name = uploaded_file.name
            date_match = re.search(r'(\d+)月(\d+)号?', file_name)
            if date_match:
                month = date_match.group(1)
                day = date_match.group(2)
                date_str = f"{month}月{day}日"
            else:
                date_str = file_name.replace('.xlsx', '').replace('.xls', '')

            wb = load_workbook(uploaded_file, data_only=True)
            ws = wb.active

            header_row = None
            machine_col = None
            inspect_col = None
            defect_col = None

            for row_idx in range(1, min(20, ws.max_row + 1)):
                row_values = [cell.value for cell in ws[row_idx]]
                row_str = ' '.join([str(v) for v in row_values if v])

                if '日期' in row_str and '机台' in row_str and '送检' in row_str:
                    header_row = row_idx
                    for col_idx, cell in enumerate(ws[row_idx], 1):
                        val = str(cell.value) if cell.value else ''
                        if '机台' in val:
                            machine_col = col_idx
                        elif '送检' in val and '轴' in val:
                            inspect_col = col_idx
                        elif '不良' in val and '轴' in val:
                            defect_col = col_idx
                    break

            if header_row is None:
                header_row = 3
                machine_col = 2
                inspect_col = 5
                defect_col = 8

            file_machine_data = {}
            current_machine = None

            for row_idx in range(header_row + 1, ws.max_row + 1):
                machine_cell = ws.cell(row_idx, machine_col) if machine_col else None
                machine_val = machine_cell.value if machine_cell else None

                inspect_qty = parse_cell_value(ws.cell(row_idx, inspect_col)) if inspect_col else 0
                defect_qty = parse_cell_value(ws.cell(row_idx, defect_col)) if defect_col else 0

                if inspect_qty == 0 and defect_qty == 0:
                    continue

                if machine_val is not None and machine_val != '':
                    machine_str = str(machine_val).strip()
                    if '合计' in machine_str or '各员工' in machine_str or '工号' in machine_str:
                        break
                    machine_normalized = normalize_machine_name(machine_str)
                    if machine_normalized:
                        current_machine = machine_normalized
                else:
                    if current_machine is None:
                        continue
                    machine_normalized = current_machine

                if current_machine is None:
                    continue

                if current_machine not in file_machine_data:
                    file_machine_data[current_machine] = {'送检数': 0, '不良数': 0}

                file_machine_data[current_machine]['送检数'] += inspect_qty
                file_machine_data[current_machine]['不良数'] += defect_qty

            for machine, data in file_machine_data.items():
                if machine not in machine_data:
                    machine_data[machine] = {'送检数': 0, '不良数': 0}
                machine_data[machine]['送检数'] += data['送检数']
                machine_data[machine]['不良数'] += data['不良数']

            daily_inspect = 0
            daily_defect = 0

            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, min(5, ws.max_column + 1)):
                    cell_val = ws.cell(row_idx, col_idx).value
                    if cell_val and '合计' in str(cell_val):
                        for c in range(col_idx, ws.max_column + 1):
                            cell_val2 = ws.cell(row_idx, c).value
                            if cell_val2 and isinstance(cell_val2, (int, float)):
                                if daily_inspect == 0:
                                    daily_inspect = float(cell_val2)
                                elif daily_defect == 0:
                                    daily_defect = float(cell_val2)
                                    break
                        break
                if daily_inspect > 0:
                    break

            if daily_inspect == 0:
                for row_idx in range(1, ws.max_row + 1):
                    first_cell = ws.cell(row_idx, 1).value
                    if first_cell and '合计' in str(first_cell):
                        daily_inspect = parse_cell_value(ws.cell(row_idx, 2))
                        daily_defect = parse_cell_value(ws.cell(row_idx, 3))
                        break

            if date_str not in daily_summary:
                daily_summary[date_str] = {'送检数': 0, '不良数': 0}
            daily_summary[date_str]['送检数'] += daily_inspect
            daily_summary[date_str]['不良数'] += daily_defect

            employee_start_row = None
            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, min(10, ws.max_column + 1)):
                    cell_val = ws.cell(row_idx, col_idx).value
                    if cell_val and '各员工合格率汇总' in str(cell_val):
                        employee_start_row = row_idx + 1
                        break
                if employee_start_row:
                    break

            if employee_start_row:
                emp_id_col = None
                emp_inspect_col = None
                emp_defect_col = None

                for col_idx in range(1, ws.max_column + 1):
                    header_val = ws.cell(employee_start_row, col_idx).value
                    if header_val:
                        header_str = str(header_val)
                        if '工号' in header_str:
                            emp_id_col = col_idx
                        elif '送检' in header_str:
                            emp_inspect_col = col_idx
                        elif '退回' in header_str or '不良' in header_str:
                            emp_defect_col = col_idx

                if emp_id_col is None:
                    emp_id_col = 1
                    emp_inspect_col = 2
                    emp_defect_col = 3

                for row_idx in range(employee_start_row + 1, min(employee_start_row + 30, ws.max_row + 1)):
                    emp_cell = ws.cell(row_idx, emp_id_col)
                    if emp_cell.value:
                        emp_str = str(emp_cell.value).strip()
                        if '#' in emp_str and len(emp_str) <= 8:
                            emp_inspect = parse_cell_value(ws.cell(row_idx, emp_inspect_col))
                            emp_defect = parse_cell_value(ws.cell(row_idx, emp_defect_col))

                            if emp_inspect > 0:
                                if emp_str not in employee_data:
                                    employee_data[emp_str] = {'送检数': 0, '不良数': 0}
                                employee_data[emp_str]['送检数'] += emp_inspect
                                employee_data[emp_str]['不良数'] += emp_defect

            wb.close()

        except Exception as e:
            st.warning(f"处理文件 {uploaded_file.name} 时出错: {str(e)}")

        progress_bar.progress((i + 1) / len(uploaded_files))

    status_text.text("✅ 处理完成！")
    return daily_summary, machine_data, employee_data


def generate_all_summaries(daily_summary, machine_data, employee_data):
    daily_list = []
    total_inspect = 0
    total_defect = 0

    for date, data in daily_summary.items():
        inspect_qty = data.get('送检数', 0)
        defect_qty = data.get('不良数', 0)
        defect_rate = (defect_qty / inspect_qty * 100) if inspect_qty > 0 else 0

        total_inspect += inspect_qty
        total_defect += defect_qty

        daily_list.append({
            '日期': date,
            '送检数': int(inspect_qty),
            '不良数': int(defect_qty),
            '不合格率': f"{defect_rate:.2f}%"
        })

    def sort_by_date(x):
        nums = re.findall(r'\d+', x['日期'])
        if len(nums) >= 2:
            return int(nums[0]) * 100 + int(nums[1])
        return 999

    daily_list.sort(key=sort_by_date)

    overall_rate = (total_defect / total_inspect * 100) if total_inspect > 0 else 0
    daily_list.append({
        '日期': '合计',
        '送检数': int(total_inspect),
        '不良数': int(total_defect),
        '不合格率': f"{overall_rate:.2f}%"
    })

    df_daily = pd.DataFrame(daily_list)

    machine_list = []
    machine_total_inspect = 0
    machine_total_defect = 0

    for machine, data in machine_data.items():
        inspect_qty = data['送检数']
        defect_qty = data['不良数']
        defect_rate = (defect_qty / inspect_qty * 100) if inspect_qty > 0 else 0

        machine_total_inspect += inspect_qty
        machine_total_defect += defect_qty

        machine_list.append({
            '机台号': machine,
            '送检数': int(inspect_qty),
            '不良数': int(defect_qty),
            '不合格率': f"{defect_rate:.2f}%"
        })

    machine_list.sort(key=lambda x: extract_number(x['机台号']))

    machine_overall_rate = (machine_total_defect / machine_total_inspect * 100) if machine_total_inspect > 0 else 0
    machine_list.append({
        '机台号': '合计',
        '送检数': int(machine_total_inspect),
        '不良数': int(machine_total_defect),
        '不合格率': f"{machine_overall_rate:.2f}%"
    })

    df_machine = pd.DataFrame(machine_list)

    employee_list = []
    emp_total_inspect = 0
    emp_total_defect = 0

    for emp_id, data in employee_data.items():
        emp_inspect = data['送检数']
        emp_defect = data['不良数']
        pass_rate = ((emp_inspect - emp_defect) / emp_inspect * 100) if emp_inspect > 0 else 0

        emp_total_inspect += emp_inspect
        emp_total_defect += emp_defect

        employee_list.append({
            '工号': emp_id,
            '送检数': int(emp_inspect),
            '退回数': int(emp_defect),
            '合格率': f"{pass_rate:.2f}%"
        })

    employee_list.sort(key=lambda x: extract_number(x['工号']))

    emp_overall_rate = (
                (emp_total_inspect - emp_total_defect) / emp_total_inspect * 100) if emp_total_inspect > 0 else 0
    employee_list.append({
        '工号': '合计',
        '送检数': int(emp_total_inspect),
        '退回数': int(emp_total_defect),
        '合格率': f"{emp_overall_rate:.2f}%"
    })

    df_employee = pd.DataFrame(employee_list)

    return df_daily, df_machine, df_employee


# ========== 主程序 ==========
def main():
    # 如果是管理员模式，显示管理员面板
    if st.session_state.get("admin_mode", False):
        admin_panel()
        return

    st.title("📊 机台不良报表月度汇总系统")
    st.markdown("上传多份日报表，自动生成三类月度汇总报表")

    if "visit_logged" not in st.session_state:
        log_usage("访问系统")
        st.session_state.visit_logged = True

    with st.sidebar:
        st.header("⚙️ 系统信息")
        st.markdown("""
        **功能：**
        - ✅ 批量上传日报表
        - ✅ 生成每日汇总
        - ✅ 生成机台汇总
        - ✅ 生成员工汇总
        - ✅ 导出多sheet Excel
        """)

        st.divider()

        summary = get_stats_summary()
        if isinstance(summary, dict):
            st.caption(f"📊 今日访问: {summary['today_visits']} 次")
            st.caption(f"📈 累计汇总: {summary['total_summaries']} 次")

        st.divider()

        if st.button("🚪 退出登录", use_container_width=True):
            log_usage("退出登录")
            st.session_state.password_correct = False
            st.session_state.admin_mode = False
            st.session_state.visit_logged = False
            st.rerun()

        st.caption("© 2026 | 版本 4.2")

    uploaded_files = st.file_uploader(
        "选择日报表文件（可多选）",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        help="支持同时上传多个Excel文件"
    )

    if uploaded_files:
        st.info(f"📁 已选择 {len(uploaded_files)} 个文件")

        if st.button("🚀 开始汇总", type="primary", use_container_width=True):
            with st.spinner("正在处理中，请稍候..."):
                daily_summary, machine_data, employee_data = process_files(uploaded_files)
                df_daily, df_machine, df_employee = generate_all_summaries(
                    daily_summary, machine_data, employee_data
                )

                log_usage("汇总完成", len(uploaded_files), f"处理了{len(uploaded_files)}个文件")

                st.success(f"✅ 汇总完成！共处理 {len(uploaded_files)} 个文件")

                tab1, tab2, tab3 = st.tabs(["📅 每日汇总", "🏭 机台汇总", "👥 员工汇总"])

                with tab1:
                    st.subheader("每日总送检/不良汇总")
                    st.dataframe(df_daily, use_container_width=True, hide_index=True)

                    total_inspect = df_daily[df_daily['日期'] != '合计']['送检数'].sum()
                    total_defect = df_daily[df_daily['日期'] != '合计']['不良数'].sum()
                    rate = (total_defect / total_inspect * 100) if total_inspect > 0 else 0

                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("月度总送检数", f"{total_inspect:,}")
                    with col2:
                        st.metric("月度总不良数", f"{total_defect:,}")
                    with col3:
                        st.metric("月度总不合格率", f"{rate:.2f}%")

                with tab2:
                    st.subheader("各机台月度汇总")
                    st.dataframe(df_machine, use_container_width=True, hide_index=True)

                with tab3:
                    st.subheader("各员工月度合格率汇总")
                    st.dataframe(df_employee, use_container_width=True, hide_index=True)

                st.subheader("📥 下载报表")

                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                        df_daily.to_excel(writer, sheet_name='每日汇总', index=False)
                        df_machine.to_excel(writer, sheet_name='机台汇总', index=False)
                        df_employee.to_excel(writer, sheet_name='员工汇总', index=False)

                    with open(tmp.name, 'rb') as f:
                        excel_data = f.read()

                    st.download_button(
                        label="📊 下载完整Excel报表",
                        data=excel_data,
                        file_name=f"3月不良汇总_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                try:
                    os.unlink(tmp.name)
                except:
                    pass

    else:
        st.markdown("""
        ### 📁 使用说明
        1. 点击上方按钮选择所有日报表文件
        2. 点击"开始汇总"按钮
        3. 查看三类汇总报表并下载

        ### 📊 输出结果
        - **每日汇总**：每天的总送检数、不良数、不合格率
        - **机台汇总**：各机台整月的送检数、不良数、不合格率
        - **员工汇总**：各员工整月的送检数、退回数、合格率
        """)


# ========== 入口 ==========
if check_password():
    main()